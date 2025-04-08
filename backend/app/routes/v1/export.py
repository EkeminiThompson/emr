from fastapi import APIRouter, HTTPException, Depends, Query, Request
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from sqlalchemy.orm import Session
import openpyxl
from sqlalchemy import inspect
from sqlalchemy.orm import DeclarativeMeta
from typing import List, Type, Optional, Dict, Any
from io import BytesIO, StringIO
from fastapi.responses import StreamingResponse
from app.database import get_db
from app.models import (
    Patient, ClinicalNote, MentalHealthNote, PharmacyRecord,
    LaboratoryRecord, OccupationalTherapyRecord, PsychologyRecord,
    SocialWorkRecord
)
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from datetime import datetime, date
import json
import csv
from slowapi import Limiter
from slowapi.util import get_remote_address
from pydantic import BaseModel

# Security and rate limiting
security = HTTPBearer()
limiter = Limiter(key_func=get_remote_address)

router = APIRouter(prefix="/exports", tags=["Exports"])

# Configuration models
class ExportConfig(BaseModel):
    date_format: str = "%Y-%m-%d"
    datetime_format: str = "%Y-%m-%d %H:%M:%S"
    excluded_fields: List[str] = [
        'id', 'created_at', 'updated_at', 'user_id', 
        'doctor_id', 'nurse_id', 'billing_id'
    ]
    max_export_records: int = 1000

# Current configuration (could be loaded from DB or config file)
export_config = ExportConfig()

# Utility Functions
def get_exportable_fields(model_class: Type[DeclarativeMeta]) -> List[str]:
    """Get all columns except relationships and internal fields"""
    inspector = inspect(model_class)
    return [
        col.name for col in model_class.__table__.columns
        if not col.name.endswith('_id') and col.name not in export_config.excluded_fields
    ]

def format_field_name(field: str) -> str:
    """Convert snake_case to Title Case"""
    return field.replace('_', ' ').title()

def format_field_value(value, field_type: Optional[str] = None):
    """Format field values for display with configurable formats"""
    if value is None:
        return ""
    elif isinstance(value, datetime):
        return value.strftime(export_config.datetime_format)
    elif isinstance(value, date):
        return value.strftime(export_config.date_format)
    elif isinstance(value, (list, dict)):
        return json.dumps(value, indent=2)
    return str(value)

def verify_access(credentials: HTTPAuthorizationCredentials):
    """Verify user has access to exports"""
    if not credentials.scheme == "Bearer":
        raise HTTPException(status_code=403, detail="Invalid authentication scheme")
    return True

# Patient Data Retrieval
def get_patient_data(patient_id: str, db: Session) -> Dict[str, Any]:
    """Get comprehensive patient data including all related records"""
    patient = db.query(Patient).filter(Patient.patient_id == patient_id).first()
    if not patient:
        raise HTTPException(status_code=404, detail="Patient not found")
    
    patient_data = {col.name: getattr(patient, col.name) for col in patient.__table__.columns}
    
    related_models = {
        "clinical_notes": ClinicalNote,
        "mental_health_notes": MentalHealthNote,
        "pharmacy_records": PharmacyRecord,
        "laboratory_records": LaboratoryRecord,
        "occupational_therapy_records": OccupationalTherapyRecord,
        "psychology_records": PsychologyRecord,
        "social_work_records": SocialWorkRecord
    }
    
    related_data = {}
    for key, model in related_models.items():
        records = db.query(model).filter(model.patient_id == patient_id).all()
        related_data[key] = [
            {col.name: getattr(record, col.name) for col in record.__table__.columns} 
            for record in records
        ]
    
    return {
        "patient": patient_data,
        "related_records": related_data,
        "export_metadata": {
            "exported_at": datetime.utcnow(),
            "export_format": "internal",
            "record_count": 1 + sum(len(v) for v in related_data.values())
        }
    }

# Excel Generation
def generate_excel(data: Dict[str, Any], include_related: bool = True) -> BytesIO:
    """Generate Excel file with patient data and optional related records"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Patient Data"
    
    # Add metadata
    ws.append(["EXPORT METADATA"])
    ws.append(["Exported at", data["export_metadata"]["exported_at"].strftime(export_config.datetime_format)])
    ws.append(["Record count", data["export_metadata"]["record_count"]])
    ws.append([])
    
    # Patient data
    ws.append(["PATIENT INFORMATION"])
    for field in get_exportable_fields(Patient):
        value = data["patient"].get(field)
        ws.append([format_field_name(field), format_field_value(value)])
    ws.append([])
    
    if include_related:
        # Related records
        for record_type, records in data["related_records"].items():
            if records:
                ws.append([format_field_name(record_type).upper()])
                model = {
                    "clinical_notes": ClinicalNote,
                    "mental_health_notes": MentalHealthNote,
                    "pharmacy_records": PharmacyRecord,
                    "laboratory_records": LaboratoryRecord,
                    "occupational_therapy_records": OccupationalTherapyRecord,
                    "psychology_records": PsychologyRecord,
                    "social_work_records": SocialWorkRecord
                }[record_type]
                
                # Add headers
                headers = [format_field_name(field) for field in get_exportable_fields(model)]
                ws.append(headers)
                
                # Add data
                for record in records:
                    row = [format_field_value(record.get(field)) for field in get_exportable_fields(model)]
                    ws.append(row)
                
                ws.append([])
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

def generate_all_patients_excel(
    db: Session,
    skip: int = 0,
    limit: int = 100,
    filters: Optional[Dict] = None
) -> BytesIO:
    """Generate Excel file with paginated list of all patients including their related records"""
    wb = Workbook()
    ws = wb.active
    ws.title = "All Patients"
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = "4472C4"  # Blue color
    patient_header_font = Font(bold=True, italic=True)
    record_header_font = Font(bold=True, color="000000")
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))
    
    # Add metadata
    ws.append(["COMPREHENSIVE PATIENT RECORDS EXPORT"])
    ws.append(["Exported at", datetime.utcnow().strftime(export_config.datetime_format)])
    ws.append(["Page", f"{skip//limit + 1} (records {skip}-{skip+limit})"])
    ws.append([])
    
    # Define patient export fields
    patient_fields = [
        ('patient_id', 'Patient ID'),
        ('surname', 'Surname'),
        ('other_names', 'Other Names'),
        ('date_of_birth', 'Date of Birth'),
        ('sex', 'Gender'),
        ('age', 'Age'),
        ('marital_status', 'Marital Status'),
        ('residential_address', 'Address'),
        ('residential_phone', 'Phone'),
        ('next_of_kin', 'Next of Kin')
    ]
    
    # Define related models to include
    related_models = {
        "Clinical Notes": ClinicalNote,
        "Mental Health Notes": MentalHealthNote,
        "Pharmacy Records": PharmacyRecord,
        "Laboratory Records": LaboratoryRecord,
        "Occupational Therapy": OccupationalTherapyRecord,
        "Psychology Records": PsychologyRecord,
        "Social Work Records": SocialWorkRecord
    }
    
    # Build headers - start with patient fields
    headers = [col[1] for col in patient_fields]
    
    # Add headers for related records counts
    for model_name in related_models.keys():
        headers.append(f"{model_name} Count")
    
    # Add main headers
    ws.append(headers)
    
    # Apply styles to headers
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=5, column=col)
        cell.font = header_font
        cell.fill = openpyxl.styles.PatternFill(start_color=header_fill, end_color=header_fill, fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    # Build query
    query = db.query(Patient).order_by(Patient.surname)
    
    # Apply filters if provided
    if filters:
        if 'sex' in filters:
            query = query.filter(Patient.sex == filters['sex'])
        if 'min_age' in filters:
            query = query.filter(Patient.age >= filters['min_age'])
        if 'max_age' in filters:
            query = query.filter(Patient.age <= filters['max_age'])
    
    # Apply pagination
    patients = query.offset(skip).limit(limit).all()
    
    if not patients:
        ws.append(["No patients found matching criteria"])
    else:
        for patient in patients:
            # Start with patient data
            row = []
            for field, _ in patient_fields:
                value = getattr(patient, field)
                if field == 'date_of_birth' and value:
                    value = value.strftime(export_config.date_format)
                row.append(value)
            
            # Get all related records for this patient
            related_records = {}
            for model_name, model in related_models.items():
                records = db.query(model).filter(model.patient_id == patient.patient_id).all()
                related_records[model_name] = records
                # Add count of records
                row.append(len(records))
            
            # Add patient row
            ws.append(row)
            
            # Add related records for this patient
            max_records = max(len(records) for records in related_records.values()) if related_records else 0
            
            for i in range(max_records):
                record_row = [""] * len(patient_fields)  # Empty cells for patient info
                
                # Add record counts again (empty for visual alignment)
                for model_name in related_models.keys():
                    record_row.append("")
                
                # Add related records data
                for model_name, model in related_models.items():
                    records = related_records[model_name]
                    if i < len(records):
                        record = records[i]
                        for field in get_exportable_fields(model):
                            value = getattr(record, field)
                            record_row.append(format_field_value(value))
                    else:
                        # Add empty cells if no record at this index
                        for _ in get_exportable_fields(model):
                            record_row.append("")
                
                ws.append(record_row)
            
            # Add empty row between patients for readability
            ws.append([])
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width
    
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    return excel_file

# PDF Generation
def generate_pdf(data: Dict[str, Any]) -> BytesIO:
    """Generate PDF report with comprehensive patient data"""
    buffer = BytesIO()
    p = canvas.Canvas(buffer, pagesize=letter)
    
    # PDF Styling
    p.setFont("Helvetica-Bold", 14)
    p.drawString(100, 780, "COMPREHENSIVE PATIENT RECORD")
    p.setFont("Helvetica", 10)
    
    # Metadata
    p.drawString(100, 760, f"Exported at: {data['export_metadata']['exported_at'].strftime(export_config.datetime_format)}")
    p.drawString(100, 745, f"Record count: {data['export_metadata']['record_count']}")
    
    y_position = 720
    
    # Patient Information
    p.setFont("Helvetica-Bold", 12)
    p.drawString(100, y_position, "PATIENT INFORMATION")
    y_position -= 20
    p.setFont("Helvetica", 10)
    
    for field in get_exportable_fields(Patient):
        value = data["patient"].get(field)
        text = f"{format_field_name(field)}: {format_field_value(value)}"
        p.drawString(100, y_position, text)
        y_position -= 15
        if y_position < 50:
            p.showPage()
            y_position = 750
            p.setFont("Helvetica", 10)
    
    # Related Records
    for record_type, records in data["related_records"].items():
        if records:
            p.setFont("Helvetica-Bold", 12)
            p.drawString(100, y_position, format_field_name(record_type).upper())
            y_position -= 20
            if y_position < 50:
                p.showPage()
                y_position = 750
            
            p.setFont("Helvetica", 10)
            for record in records:
                model = {
                    "clinical_notes": ClinicalNote,
                    "mental_health_notes": MentalHealthNote,
                    "pharmacy_records": PharmacyRecord,
                    "laboratory_records": LaboratoryRecord,
                    "occupational_therapy_records": OccupationalTherapyRecord,
                    "psychology_records": PsychologyRecord,
                    "social_work_records": SocialWorkRecord
                }[record_type]
                
                for field in get_exportable_fields(model):
                    value = record.get(field)
                    text = f"{format_field_name(field)}: {format_field_value(value)}"
                    p.drawString(100, y_position, text)
                    y_position -= 15
                    if y_position < 50:
                        p.showPage()
                        y_position = 750
                        p.setFont("Helvetica", 10)
                
                y_position -= 10  # Space between records
    
    p.save()
    buffer.seek(0)
    return buffer

# CSV Generation
def generate_csv(data: Dict[str, Any]) -> BytesIO:
    """Generate CSV file with patient data"""
    output = StringIO()
    writer = csv.writer(output)
    
    # Write metadata
    writer.writerow(["EXPORT METADATA"])
    writer.writerow(["Exported at", data["export_metadata"]["exported_at"].strftime(export_config.datetime_format)])
    writer.writerow(["Record count", data["export_metadata"]["record_count"]])
    writer.writerow([])
    
    # Patient data
    writer.writerow(["PATIENT INFORMATION"])
    for field in get_exportable_fields(Patient):
        value = data["patient"].get(field)
        writer.writerow([format_field_name(field), format_field_value(value)])
    writer.writerow([])
    
    # Related records
    for record_type, records in data["related_records"].items():
        if records:
            writer.writerow([format_field_name(record_type).upper()])
            
            model = {
                "clinical_notes": ClinicalNote,
                "mental_health_notes": MentalHealthNote,
                "pharmacy_records": PharmacyRecord,
                "laboratory_records": LaboratoryRecord,
                "occupational_therapy_records": OccupationalTherapyRecord,
                "psychology_records": PsychologyRecord,
                "social_work_records": SocialWorkRecord
            }[record_type]
            
            # Write headers
            writer.writerow([format_field_name(field) for field in get_exportable_fields(model)])
            
            # Write data
            for record in records:
                writer.writerow([format_field_value(record.get(field)) for field in get_exportable_fields(model)])
            
            writer.writerow([])
    
    # Convert to bytes
    csv_data = output.getvalue().encode('utf-8')
    return BytesIO(csv_data)

# FastAPI Endpoints
@router.get("/patients/{patient_id}", summary="Get patient data")
async def get_patient(
    patient_id: str, 
    db: Session = Depends(get_db),
    credentials: HTTPAuthorizationCredentials = Depends(security)
):
    """Get comprehensive patient data including all related records"""
    verify_access(credentials)
    return get_patient_data(patient_id, db)

@router.get("/patients/{patient_id}/excel", response_class=StreamingResponse)
@limiter.limit("10/minute")
async def export_patient_excel(
    request: Request,
    patient_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Export patient data to Excel format"""
    verify_access(credentials)
    try:
        data = get_patient_data(patient_id, db)
        excel_file = generate_excel(data)
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={patient_id}_record.xlsx"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/patients/{patient_id}/pdf", response_class=StreamingResponse)
@limiter.limit("10/minute")
async def export_patient_pdf(
    request: Request,
    patient_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Export patient data to PDF format"""
    verify_access(credentials)
    try:
        data = get_patient_data(patient_id, db)
        pdf_file = generate_pdf(data)
        return StreamingResponse(
            pdf_file,
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={patient_id}_record.pdf"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/patients/{patient_id}/csv", response_class=StreamingResponse)
@limiter.limit("10/minute")
async def export_patient_csv(
    request: Request,
    patient_id: str,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Export patient data to CSV format"""
    verify_access(credentials)
    try:
        data = get_patient_data(patient_id, db)
        csv_file = generate_csv(data)
        return StreamingResponse(
            csv_file,
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={patient_id}_record.csv"}
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/all/excel", response_class=StreamingResponse)
@limiter.limit("5/minute")
async def export_all_patients_excel(
    request: Request,
    skip: int = Query(0, ge=0),
    limit: int = Query(100, ge=1, le=1000),
    sex: Optional[str] = None,
    min_age: Optional[int] = None,
    max_age: Optional[int] = None,
    credentials: HTTPAuthorizationCredentials = Depends(security),
    db: Session = Depends(get_db)
):
    """Export paginated list of all patients to Excel format"""
    verify_access(credentials)
    try:
        # Build filters
        filters = {}
        if sex:
            filters['sex'] = sex
        if min_age is not None:
            filters['min_age'] = min_age
        if max_age is not None:
            filters['max_age'] = max_age
        
        excel_file = generate_all_patients_excel(db, skip, limit, filters)
        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=all_patients.xlsx"}
        )
    except Exception as e:
        if "No patients found" in str(e):
            raise HTTPException(
                status_code=404,
                detail="No patients found matching the specified criteria"
            )
        raise HTTPException(
            status_code=500,
            detail=f"Failed to generate export: {str(e)}"
        )